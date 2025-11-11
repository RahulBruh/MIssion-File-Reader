import pandas as pd
from openpyxl import load_workbook
import pygame
import math
import win32com.client as win32
import os

pygame.init()

screen = pygame.display.set_mode((800, 800))
clock = pygame.time.Clock()
running = True
            # x, y, width, height
jan = pygame.Rect(100, 100, 200, 50)
feb = pygame.Rect(100, 200, 200, 50)
mar = pygame.Rect(100, 300, 200, 50)
apr = pygame.Rect(100, 400, 200, 50)
may = pygame.Rect(100, 500, 200, 50)
jun = pygame.Rect(100, 600, 200, 50)
jul = pygame.Rect(100, 700, 200, 50)
aug = pygame.Rect(400, 100, 200, 50)
sept = pygame.Rect(400, 200, 200, 50)
oct = pygame.Rect(400, 300, 200, 50)
nov = pygame.Rect(400, 400, 200, 50)
dec = pygame.Rect(400, 500, 200, 50)

jan_text = pygame.font.Font(None, 36).render("January", True, (255, 255, 255))
feb_text = pygame.font.Font(None, 36).render("February", True, (255, 255, 255))
mar_text = pygame.font.Font(None, 36).render("March", True, (255, 255, 255))
apr_text = pygame.font.Font(None, 36).render("April", True, (255, 255, 255))
may_text = pygame.font.Font(None, 36).render("May", True, (255, 255, 255))
jun_text = pygame.font.Font(None, 36).render("June", True, (255, 255, 255))
jul_text = pygame.font.Font(None, 36).render("July", True, (255, 255, 255))
aug_text = pygame.font.Font(None, 36).render("August", True, (255, 255, 255))
sept_text = pygame.font.Font(None, 36).render("September", True, (255, 255, 255))
oct_text = pygame.font.Font(None, 36).render("October", True, (255, 255, 255))
nov_text = pygame.font.Font(None, 36).render("November", True, (255, 255, 255))
dec_text = pygame.font.Font(None, 36).render("December", True, (255, 255, 255))

month = None



while running:
    for event in pygame.event.get():
        screen.fill(("pink"))

        jan_button = pygame.draw.rect(screen, ("black"), jan)
        jan_rect = jan_text.get_rect(center=jan.center)
        screen.blit(jan_text, jan_rect)

        feb_button = pygame.draw.rect(screen, ("black"), feb)
        feb_rect = feb_text.get_rect(center=feb.center)
        screen.blit(feb_text, feb_rect)

        mar_button = pygame.draw.rect(screen, ("black"), mar)
        mar_rect = mar_text.get_rect(center=mar.center)
        screen.blit(mar_text, mar_rect)

        apr_button = pygame.draw.rect(screen, ("black"), apr)
        apr_rect = apr_text.get_rect(center=apr.center)
        screen.blit(apr_text, apr_rect)

        may_button = pygame.draw.rect(screen, ("black"), may)
        may_rect = may_text.get_rect(center=may.center)
        screen.blit(may_text, may_rect)

        jun_button = pygame.draw.rect(screen, ("black"), jun)
        jun_rect = jun_text.get_rect(center=jun.center)
        screen.blit(jun_text, jun_rect)

        jul_button = pygame.draw.rect(screen, ("black"), jul)
        jul_rect = jul_text.get_rect(center=jul.center)
        screen.blit(jul_text, jul_rect)
        

        aug_button = pygame.draw.rect(screen, ("black"), aug)
        aug_rect = aug_text.get_rect(center=aug.center)
        screen.blit(aug_text, aug_rect)

        sept_button = pygame.draw.rect(screen,("black"), sept)
        sept_rect = sept_text.get_rect(center=sept.center)
        screen.blit(sept_text, sept_rect)

        oct_button = pygame.draw.rect(screen, ("black"), oct)
        oct_rect = oct_text.get_rect(center=oct.center)
        screen.blit(oct_text, oct_rect)

        nov_button = pygame.draw.rect(screen, ("black"), nov)
        nov_rect = nov_text.get_rect(center=nov.center)
        screen.blit(nov_text, nov_rect)

        dec_button = pygame.draw.rect(screen, ("black"), dec)
        dec_rect = dec_text.get_rect(center=dec.center)
        screen.blit(dec_text, dec_rect)


        pygame.display.update()
        clock.tick(60)


        if event.type == pygame.QUIT:
            pygame.quit()
            running  = False
        if event.type == pygame.MOUSEBUTTONDOWN:
            if jan.collidepoint(event.pos):
                month = "January"
                running  = False
                pygame.quit()
            elif feb.collidepoint(event.pos):
                month = "February"
                running  = False
                pygame.quit()
            elif mar.collidepoint(event.pos):
                month = "March"
                running  = False
                pygame.quit()
            elif apr.collidepoint(event.pos):
                month = "April"
                running  = False
                pygame.quit()
            elif may.collidepoint(event.pos):
                month = "May"
                running  = False
                pygame.quit()
            elif jun.collidepoint(event.pos):
                month = "June"
                running  = False
                pygame.quit()
            elif jul.collidepoint(event.pos):
                month = "July"
                running  = False
                pygame.quit()
            elif aug.collidepoint(event.pos):
                month = "August"
                running  = False
                pygame.quit()
            elif sept.collidepoint(event.pos):
                month = "September"
                running  = False
                pygame.quit()
            elif oct.collidepoint(event.pos):
                month = "October"
                running  = False
                pygame.quit()
            elif nov.collidepoint(event.pos):
                month = "November"
                running  = False
                pygame.quit()
            elif dec.collidepoint(event.pos):
                month = "December"
                running  = False
                pygame.quit()
if month is None:
    print("No month selected. Exiting.")
    exit()
   
    


months_column = {
    "January": 9,
    "February": 10,
    "March": 11,
    "April": 12,
    "May": 13,
    "June": 14,
    "July": 15,
    "August": 4,
    "September": 5,
    "October": 6,
    "November": 7,
    "December": 8
}

BUD_PATH = "Default_2025_Budget_Import.xlsx"
BUD_SHEET = "Default"
BAL_PATH = "Balance_Sheet_Default_2025.xlsx"
BAL_SHEET = "August"

def _norm_label(v):
    return str(v).strip().casefold()

def legitimize_excel(file_path):
    """
    Forces Excel to recalculate all formulas by opening and saving the file 
    through the Excel application.
    """
    print(f"Legitimizing Excel file: {file_path}")
    # Convert to absolute path
    abs_path = os.path.abspath(file_path)
    
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False  # Suppress any Excel alerts
    
    wb = excel.Workbooks.Open(abs_path)
    
    # Force full calculation
    excel.CalculateFullRebuild()
    
    # Save it again, fully rebuilt
    wb.Save()  # Use Save() instead of SaveAs() to keep the same file
    wb.Close(SaveChanges=True)
    excel.Quit()
    
    print(f"Successfully legitimized: {file_path}")

def first_four(number, rest_number, month):
    df = pd.read_excel(BUD_PATH, header=0)
    wb = load_workbook(BUD_PATH, data_only=True)
    ws = wb[BUD_SHEET]

    for idx, r in df.iloc[0:30].iterrows():
        label = _norm_label(r.iloc[2])
        if label in {
            "comp budget allocation",
            "total personnel",
            "total sub total compensation",
            "total expense",          
            "total expenses",
            "budget allocation",
            "total ministry costs",
            "total ministry cost",
            "total net surplus(deficit)",
            "total net surplus (deficit)"


        }:
            for col in range(months_column[month], 16):
                if col == months_column[month]:
                    ws.cell(row=idx + 2, column=months_column[month]).value = number
                else:
                    ws.cell(row=idx + 2, column=col).value = rest_number
            ws.cell(row=idx + 2, column=16).value = 0
            ws.cell(row=idx + 2, column=16).value = math.ceil(sum(ws.cell(row=idx + 2, column=c).value or 0 for c in range(4, 17)))



    wb.save(BUD_PATH)

def two_seven(number, start_excel_row, end_excel_row, rest_number, month):
    """Apply within [start, end) where start/end are Excel row numbers (1-based)."""
    start_idx = start_excel_row - 2  
    end_idx   = end_excel_row   - 2
    df = pd.read_excel(BUD_PATH, header=0)
    wb = load_workbook(BUD_PATH, data_only=True)
    ws = wb[BUD_SHEET]

    for idx, r in df.iloc[start_idx:end_idx].iterrows():
        label = _norm_label(r.iloc[2])
        if label in {
            "comp budget allocation",
            "total expenses",
            "total sub total compensation",
            "total personnel",
            "total expense",
            "budget allocation",
            "total ministry costs",
            "total ministry cost",
            "total expenses",
            "total net surplus(deficit)"
            "total net surplus (deficit)"

        }:
            for col in range(months_column[month], 16):
                if col == months_column[month]:
                    ws.cell(row=idx + 2, column=months_column[month]).value = number
                else:
                    ws.cell(row=idx + 2, column=col).value = rest_number
            ws.cell(row=idx + 2, column=16).value = 0
            ws.cell(row=idx + 2, column=16).value = math.ceil(sum(ws.cell(row=idx + 2, column=c).value or 0 for c in range(4, 17)))


    wb.save(BUD_PATH)

def onward(number, start_excel_row, end_excel_row, rest_number, month):
    """Apply within [start, end) for the onward segments."""
    start_idx = start_excel_row - 2
    end_idx   = end_excel_row   - 2
    df = pd.read_excel(BUD_PATH, header=0)
    wb = load_workbook(BUD_PATH, data_only=True)
    ws = wb[BUD_SHEET]

    for idx, r in df.iloc[start_idx:end_idx].iterrows():
        label = _norm_label(r.iloc[2])
        if label in {
            "budget allocation",
            "total ministry cost",     
            "total ministry costs",
            "total expenses",
            "total net surplus(deficit)",
            "total net surplus (deficit)",
            "total personnel",
            "total sub total compensation",
            "total expense",
            "comp budget allocation"
        }:
            for col in range(months_column[month], 16):
                if col == months_column[month]:
                    ws.cell(row=idx + 2, column=months_column[month]).value = number
                else:
                    ws.cell(row=idx + 2, column=col).value = rest_number
            ws.cell(row=idx + 2, column=16).value = 0
            ws.cell(row=idx + 2, column=16).value = math.ceil(sum(ws.cell(row=idx + 2, column=c).value or 0 for c in range(4, 17)))

            

    wb.save(BUD_PATH)

# ----------------- Driver -----------------


df_budget = pd.read_excel(BUD_PATH, header=0)
df_balance = pd.read_excel(BAL_PATH, header=0)

wb_bal = load_workbook(BAL_PATH, data_only=True)
ws_bal = wb_bal[BAL_SHEET]


i = 0
x = 0


bruh = [30, 50, 70, 90, 110, 130, 156, 175, 245]  # Now has 9 ranges


onward_array = [245, 270, 321, 395, 472, 548, 676, 727, 780, 851, 924, 990, 1067, 1142, 1214, 1285, 1360, 1433, 1512, 1585, 1660, 1735, 1816, 1890, 1955, 2029, 2110, 2191, 2272, 2386, 2479, 2564] 


for idx, r in df_balance.iterrows():
    cell = ws_bal.cell(row=idx + 2, column=4).value
    rest_number = ws_bal.cell(row=idx + 2, column=7).value
    if not isinstance(cell, (float, int)):
        continue  

    number = cell

    if i == 0:
        first_four(number, rest_number, month)
        i = 1
        x = 0
        continue

    elif i == 1:
        if x < len(bruh) - 1:
            two_seven(number, bruh[x], bruh[x + 1], rest_number, month)
            x += 1
            if x >= len(bruh) - 1:
                i = 2
                x = 0
        continue

    elif i == 2:
        if x < len(onward_array) - 1:
            onward(number, onward_array[x], onward_array[x + 1], rest_number, month)
            x += 1
            if x >= len(onward_array) - 1:
                break
        continue

# After all updates are complete, legitimize the Excel file to recalculate formulas
print("\nAll updates complete. Now recalculating formulas...")
legitimize_excel(BUD_PATH)
print("Process complete!")
