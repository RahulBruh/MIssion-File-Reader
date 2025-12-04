import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def dropdown():
    BUD_PATH = "New+Check+Request+Form.xlsx"
    CARD_SHEET = "Credit Card Report"
    DATA_SHEET = "Data"
    
    # Read the sheets
    cf = pd.read_excel(BUD_PATH, header=0, sheet_name=CARD_SHEET)
    df = pd.read_excel(BUD_PATH, header=0, sheet_name=DATA_SHEET)

    wb = load_workbook(BUD_PATH)
    ws = wb[CARD_SHEET]

    # Process rows 13-29 (Excel rows, which are iloc 12:29 with 0-based index)
    for excel_row in range(13, 30):  # 13 to 29 inclusive
        # Read the label from column D (index 3) in the current row
        pandas_row = excel_row - 2  # Adjust for header row
        
        if pandas_row >= len(cf):
            continue
            
        label = cf.iloc[pandas_row, 3]  # Column D
        
        print(f"Row {excel_row}: Label = '{label}'")  # Debug
        
        if pd.notna(label) and str(label).strip():
            # Look for this label in the Data sheet's first row (header)
            if label in df.columns:
                print(f"  Found '{label}' in Data sheet columns")  # Debug
                
                # Get all non-empty items from row 2 onwards (iloc[1:])
                dropdown = []
                for item in df[label].dropna().iloc[1:]:  # Skip header, drop NaN
                    dropdown.append(str(item).strip())
                
                print(f"  Dropdown items: {dropdown[:5]}...")  # Debug first 5
                
                if dropdown:
                    # Clear the cell in column E
                    ws.cell(row=excel_row, column=5).value = None
                    
                    # Create dropdown with proper formula
                    dropdown_formula = '"{}"'.format(",".join(dropdown))
                    dv = DataValidation(
                        type="list", 
                        formula1=dropdown_formula,
                        allow_blank=True,
                        showDropDown=True  # Explicitly show dropdown
                    )
                    dv.error = 'Invalid selection'
                    dv.errorTitle = 'Invalid Entry'
                    
                    cell_ref = f"E{excel_row}"
                    dv.add(cell_ref)
                    ws.add_data_validation(dv)
                    print(f"  Added dropdown to {cell_ref}")  # Debug
            else:
                print(f"  '{label}' NOT found in Data sheet columns")  # Debug
                print(f"  Available columns: {list(df.columns)[:10]}")  # Debug
    
    wb.save(BUD_PATH)
    print("File saved successfully!")

dropdown()
